import sched, time, datetime
import requests, sys
import asyncio, datetime, sys
import requests
from random import seed
from random import randint
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import time
import json, os, xlsxwriter, datetime, timedelta, boto3, sys
from openpyxl import load_workbook


url_engine = "http://augusto-accorsi-webapp-elb-1925434936.sa-east-1.elb.amazonaws.com:3001/engine/mand?max_iter=999&width=999&height=999"
url_db = 'http://augusto-accorsi-webapp-elb-1925434936.sa-east-1.elb.amazonaws.com:3001/database/save'
count = 0
minutes = 20

def save_call_db(delta, response):
    workbook = load_workbook(filename = 'dataset\\calls\\db.xlsx')
    worksheet = workbook['Sheet1']
    
    newRowLocation = worksheet.max_row +1

    #worksheet.cell(column=1,row=newRowLocation, value=start)
    #worksheet.cell(column=2,row=newRowLocation, value=end)
    worksheet.cell(column=3,row=newRowLocation, value=delta)
    worksheet.cell(column=4,row=newRowLocation, value=response)

    workbook.save(filename = 'dataset\\calls\\db.xlsx')
    workbook.close()

def save_call_engine(delta, response):
    workbook = load_workbook(filename = 'dataset\\calls\\engine.xlsx')
    worksheet = workbook['Sheet1']
    
    newRowLocation = worksheet.max_row +1

    #worksheet.cell(column=1,row=newRowLocation, value=start)
    #worksheet.cell(column=2,row=newRowLocation, value=end)
    worksheet.cell(column=3,row=newRowLocation, value=delta)
    worksheet.cell(column=4,row=newRowLocation, value=response)

    workbook.save(filename = 'dataset\\calls\\engine.xlsx')
    workbook.close()

def engine(count):
    t_end = time.time() + 60 * minutes
    while time.time() < t_end:
        session = requests.Session()
        session.trust_env = False
        res = session.post(url_engine)
        print("Call "+str(count+1)+": at "+str(datetime.datetime.now().strftime('%H:%M:%S'))+" : "+ str(res.status_code)+" : "+str(res.elapsed.total_seconds()))
        if res.status_code == 201:
            save_call_engine(res.elapsed.total_seconds(), res.status_code)
        
        #s.enter(1, 1, engine, (sc,count+1))
    print("Waiting for "+str(minutes)+" minutes at "+str(datetime.datetime.now().strftime('%H:%M:%S')))
    time.sleep(60*minutes)
    engine(count)

def db(count):
    t_end = time.time() + 60 * minutes
    while time.time() < t_end:
        count +=1
        session = requests.Session()
        session.trust_env = False
        res = session.post(url=url_db, files={'image': ('file.PNG', 'image.png', 'image/png')})
        print("Call "+str(count)+": at "+str(datetime.datetime.now().strftime('%H:%M:%S'))+" : "+ str(res.status_code)+" : "+str(res.elapsed.total_seconds()))
        if res.status_code == 201:
            save_call_db(res.elapsed.total_seconds(), res.status_code)
        
        #s.enter(1, 1, db, (sc,count+1))
    print("Waiting for "+str(minutes)+" minutes at "+str(datetime.datetime.now().strftime('%H:%M:%S')))
    time.sleep(60*minutes)
    db(count)


if not os.path.isfile('dataset\\calls\\db.xlsx'):
    workbook = xlsxwriter.Workbook('dataset\\calls\\db.xlsx')
    worksheet = workbook.add_worksheet()
    format = workbook.add_format({'num_format': 'dd/mm/yy hh:mm'})
    worksheet.write('A1', 'start', format)
    worksheet.write('B1', 'end')
    worksheet.write('C1', 'delta')
    worksheet.write('D1', 'response code')
    workbook.close()

        
if not os.path.isfile('dataset\\calls\\engine.xlsx'):
    workbook = xlsxwriter.Workbook('dataset\\calls\\engine.xlsx')
    worksheet = workbook.add_worksheet()
    format = workbook.add_format({'num_format': 'dd/mm/yy hh:mm'})
    worksheet.write('A1', 'start', format)
    worksheet.write('B1', 'end')
    worksheet.write('C1', 'delta')
    worksheet.write('D1', 'response code')
    workbook.close()

if sys.argv[1] == "engine":
    print("Waiting for "+str(minutes)+" minutes at "+str(datetime.datetime.now().strftime('%H:%M:%S')))
    time.sleep(60*minutes)
    print("Calling url engine...")
    engine(count)

if sys.argv[1] == "db":
    print("Waiting for "+str(minutes)+" minutes at "+str(datetime.datetime.now().strftime('%H:%M:%S')))
    time.sleep(60*minutes)
    print("Calling url db...")
    db(count)