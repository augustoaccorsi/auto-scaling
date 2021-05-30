
import asyncio, datetime, sys
import requests
from random import seed
from random import randint
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.retry import Retry
import time
import json, os, xlsxwriter, datetime, timedelta, boto3, sys
from openpyxl import load_workbook

class Run:

    def __init__(self):        
        if not os.path.isfile('dataset\\calls\\db.xlsx'):
            workbook = xlsxwriter.Workbook('dataset\\calls\\db.xlsx')
            worksheet = workbook.add_worksheet()
            format = workbook.add_format({'num_format': 'dd/mm/yy hh:mm'})
            worksheet.write('A1', 'start', format)
            worksheet.write('B1', 'end')
            worksheet.write('C1', 'delta')
            worksheet.write('D1', 'response code')
            workbook.close()

        
        if not os.path.isfile('dataset\\calls\\engine.xlsx'):
            workbook = xlsxwriter.Workbook('dataset\\calls\\engine.xlsx')
            worksheet = workbook.add_worksheet()
            format = workbook.add_format({'num_format': 'dd/mm/yy hh:mm'})
            worksheet.write('A1', 'start', format)
            worksheet.write('B1', 'end')
            worksheet.write('C1', 'delta')
            worksheet.write('D1', 'response code')
            workbook.close()
    
    #def call_request_get(self, count):
    #    session = requests.Session()
    #    session.trust_env = False
    #    res = session.get(self._url) 
    #    print("Call " + str(count+1) + " on "+self._url+" : " + str(res.status_code))
    #    return res.status_code

    def save_call_db(self, start, end, delta, response):
        workbook = load_workbook(filename = 'dataset\\calls\\db.xlsx')
        worksheet = workbook['Sheet1']
        
        newRowLocation = worksheet.max_row +1

        worksheet.cell(column=1,row=newRowLocation, value=start)
        worksheet.cell(column=2,row=newRowLocation, value=end)
        worksheet.cell(column=3,row=newRowLocation, value=delta)
        worksheet.cell(column=4,row=newRowLocation, value=response)

        workbook.save(filename = 'dataset\\calls\\db.xlsx')
        workbook.close()

    def save_call_engine(self, start, end, delta, response):
        workbook = load_workbook(filename = 'dataset\\calls\\engine.xlsx')
        worksheet = workbook['Sheet1']
        
        newRowLocation = worksheet.max_row +1

        worksheet.cell(column=1,row=newRowLocation, value=start)
        worksheet.cell(column=2,row=newRowLocation, value=end)
        worksheet.cell(column=3,row=newRowLocation, value=delta)
        worksheet.cell(column=4,row=newRowLocation, value=response)

        workbook.save(filename = 'dataset\\calls\\engine.xlsx')
        workbook.close()

    def call_request_post(self, count, x, y, z):
        session = requests.Session()
        session.trust_env = False
        url = "http://augusto-accorsi-webapp-elb-1925434936.sa-east-1.elb.amazonaws.com:3001/"
        #url = "http://augusto-accorsi-engine-elb-1884692720.sa-east-1.elb.amazonaws.com:5000/"
        path = "engine/mand?max_iter=&1&width=&2&height=&3"
        path = path.replace("&1", x)
        path = path.replace("&2", y)
        path = path.replace("&3", z)
        start = datetime.datetime.now()
        res = session.post(url+path)    
        end = datetime.datetime.now()
        if res.status_code == 201:
            self.save_call_engine(start, end, res.elapsed.total_seconds(), res.status_code)
        print("Call " + str(count+1) + " on "+path+" on "+str(datetime.datetime.now().strftime('%m/%d/%Y %H:%M:%S'))+" : "+ str(res.status_code))
        return res.status_code

    def call_database(self, count):
        url = 'http://augusto-accorsi-webapp-elb-1925434936.sa-east-1.elb.amazonaws.com:3001/database/save'
        session = requests.Session()
        session.trust_env = False
        start = datetime.datetime.now()
        res = session.post(url=url, files={'image': ('file.PNG', 'image.png', 'image/png')})
        end = datetime.datetime.now()
        if res.status_code == 201:
            self.save_call_db(start, end, res.elapsed.total_seconds(), res.status_code)
        print("Call " + str(count+1) + " on /database/save on "+str(datetime.datetime.now().strftime('%m/%d/%Y %H:%M:%S'))+" : "+ str(res.status_code))
        return res.status_code


        
    async def call_url(self, type):
        #await asyncio.sleep(600)
        count = 0
        
        print("sleeping for 20 minutes")
        await asyncio.sleep(20*60)
        print("woke up")

        while True:
            minutes = 20  #randint(5, 20)
            sleep = (20*60)#(randint(10, 20) * 60)
            print("calls: "+str(minutes))
            print("sleep: "+str(sleep/60))
            exit_code = 0
            finish_time = datetime.datetime.now() + datetime.timedelta(minutes=minutes)
            while datetime.datetime.now() < finish_time:
                x = "999"#str(randint(600, 800))
                y = "999"#str(randint(600, 800))
                z = "999"#str(randint(600, 800))
                if type == "engine":
                    status_code = self.call_request_post(count, x, y, z)   
                if type == "database":
                    status_code = self.call_database(count)
                if status_code != 201 and status_code != 200:
                    exit_code+=1
                if exit_code == 15:
                    exit_code = 0
                    break
                count+=1
            print("----------------------------------")
            await asyncio.sleep(sleep)

if __name__ == '__main__':
    run = Run()
    print("Starting Event Loop")
    loop = asyncio.get_event_loop()
    try:
        print("----------------------------------")
        try:
            if sys.argv[1] == "engine":
                asyncio.ensure_future(run.call_url("engine"))
            if sys.argv[1] == "database":
                asyncio.ensure_future(run.call_url("database"))
            if sys.argv[1] == "sleep":
                print("sleep for 5")
                time.sleep(300) #5 minutes
                print("just woke up")
                asyncio.ensure_future(run.call_url("post"))
        except:
            try:
                for i in range(int(sys.argv[1])):
                    if sys.argv[2] == "post":
                        run.call_request_post(i, 950, 950, 950)
                    else:
                        run.call_request_get(i)
            except:
                asyncio.ensure_future(run.call_url("post"))
            pass
        loop.run_forever()
    except KeyboardInterrupt:
        pass
    finally:
        print("Closing Event Loop")
        loop.close()

'''
class Run:

    def __init__(self):        
        #self._urlPost = "http://augusto-accorsi-engine-elb-1884692720.sa-east-1.elb.amazonaws.com:5000/engine/mand?max_iter=850&width=850&height=850"
        self._urlPost = "http://augusto-accorsi-engine-elb-1884692720.sa-east-1.elb.amazonaws.com:5000/engine/mand?max_iter=1000&width=1000&height=1000"
        self._urlGet = "http://augusto-accorsi-engine-elb-1884692720.sa-east-1.elb.amazonaws.com:5000"

    def call_request_get(self, count):
        res = requests.get(self._urlGet)
        print("Call " + str(count+1) + " on "+self._urlGet+" : " + str(res.status_code))
    def call_request_post(self, count):
        res = requests.post(self._urlPost)
        print("Call " + str(count+1) + " on "+self._urlPost+" : " + str(res.status_code))
        
    async def call_url(self, type, minutes):
        count = 0
        while True:
            finish_time = datetime.datetime.now() + datetime.timedelta(minutes=minutes)
            while datetime.datetime.now() < finish_time:
                if type == "post":
                    self.call_request_post(count)  
                if type == "get":
                    self.call_request_get(count)
                count+=1
            print("----------------------------------")
            await asyncio.sleep(1200)
if __name__ == '__main__':
    run = Run()
    print("Starting Event Loop")
    loop = asyncio.get_event_loop()
    try:
        print("----------------------------------")
        asyncio.ensure_future(run.call_url("post", 10))
        loop.run_forever()
    except KeyboardInterrupt:
        pass
    finally:
        print("Closing Event Loop")
        loop.close()    
'''