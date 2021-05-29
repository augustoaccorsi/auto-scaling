from time import time
from autoscalinggroup import Instance
from timeseries import Timeseries
import threading, queue, datetime
from openpyxl import load_workbook
from microservice import Microservice
import pandas as pd
import numpy as np
import statistics, math

class Autoscaling:
    def __init__(self, instances, autoScalingGroup, autoScalingClient):
        self._instances = instances
        self._auto_scaling_group = autoScalingGroup
        self._autoScalingClient = autoScalingClient
        self._terminated = [] 

        self._PROACTIVE_DIFF = 45
        self._ACCURACY = 70
        self._SCALE_UP = 70
        self._SCALE_DOWN = 30

        self._AUTOSCALING_TRIGGER = 3

        self._CPU_UPPER_TRESHOLD = None
        self._CPU_LOWER_TRESHOLD = None

        self._NETWORK_UPPER_TRESHOLD = None
        self._NETWORK_LOWER_TRESHOLD = None

        self._CPU_UPPER_TRESHOLD = None
        self._CPU_LOWER_TRESHOLD = None

        self._NETWORK_UPPER_TRESHOLD = None
        self._NETWORK_LOWER_TRESHOLD = None
        
        self._cooldown = False
        
        self.set_thresholds()
        
    def calculate_threasholds(self, values):
        upper =  np.percentile(np.sort(values), 90)

        if upper - int(upper) >= 0.5:
            upper = round(upper)

        #lower = math.floor(statistics.median(np.sort(values)))
        lower = statistics.median(np.sort(values))
        mean = np.mean(np.sort(values))



        return upper, lower

    def get_dataset_data(self, dataset):
        data_xls = pd.read_excel('dataset\\'+self._auto_scaling_group.getAutoScalingGroupName()+'\\'+dataset+'.xlsx', 'Sheet1', dtype=str, index_col=None)
        data_xls.to_csv('dataset\\'+self._auto_scaling_group.getAutoScalingGroupName()+'\\'+dataset+'.csv', encoding='utf-8', index=False) 
        df=pd.read_csv('dataset\\'+self._auto_scaling_group.getAutoScalingGroupName()+'\\'+dataset+'.csv')

        date, value = df.date.to_list(), df.value.to_list()
        
        try:
            last_read = datetime.datetime.strptime(date[len(date)-1], '%m/%d/%Y %H:%M:%S')
        except:
            last_read = datetime.datetime.strptime(date[len(date)-1], '%m/%d/%Y %H:%M')

        values = []

        for i in reversed(range(len(value))):
            
            try:
                aux = (last_read - datetime.datetime.strptime(date[i], '%m/%d/%Y %H:%M:%S')).total_seconds()
            except:
                aux = (last_read - datetime.datetime.strptime(date[i], '%m/%d/%Y %H:%M')).total_seconds()
            
            if aux <= 3600:
            #if aux != 0:
                values.append(value[i])
        
        return self.calculate_threasholds(values) 

    def set_thresholds(self):
        try:
            self._CPU_UPPER_TRESHOLD, self._CPU_LOWER_TRESHOLD = self.get_dataset_data('cpu')
            self._NETWORK_UPPER_TRESHOLD, self._NETWORK_LOWER_TRESHOLD =  self.get_dataset_data('network')
        except:
            self._CPU_LOWER_TRESHOLD  = 30
            self._CPU_UPPER_TRESHOLD = 70
            self._NETWORK_UPPER_TRESHOLD = 999999
            self._NETWORK_LOWER_TRESHOLD = -1

        if self._CPU_UPPER_TRESHOLD == 0:
            self._CPU_UPPER_TRESHOLD = 1
        
        print("CPU: "+str(self._CPU_UPPER_TRESHOLD)+" / "+str(self._CPU_LOWER_TRESHOLD))
        print("NET:  "+str(self._NETWORK_UPPER_TRESHOLD)+" / "+str(self._NETWORK_LOWER_TRESHOLD))

    
    def scale_up(self, instancesUp, microservice):
        if microservice._scale_up_trigger == self._AUTOSCALING_TRIGGER:
            try:
                self._autoScalingClient.set_desired_capacity(AutoScalingGroupName=self._auto_scaling_group.getAutoScalingGroupName(), DesiredCapacity=(self._auto_scaling_group.getDesiredCapacity() + instancesUp))
                print("Autoscaling Group scalled up, from "+str(self._auto_scaling_group.getDesiredCapacity())+" to "+str(self._auto_scaling_group.getDesiredCapacity() + instancesUp)+" new desired capacity: "+str(self._auto_scaling_group.getDesiredCapacity() + instancesUp))
                microservice._scale_up_trigger = 0
                self._cooldown = True
                #return True
            except:
                print("too many instances already")
        return True

    def scale_down(self, instancesDown, microservice):
        if self._auto_scaling_group.getDesiredCapacity() > 1 and microservice._scale_down_trigger == self._AUTOSCALING_TRIGGER:
            self._autoScalingClient.set_desired_capacity(AutoScalingGroupName=self._auto_scaling_group.getAutoScalingGroupName(), DesiredCapacity=(self._auto_scaling_group.getDesiredCapacity() - instancesDown))
            print("Autoscaling Group scalled down, from "+str(self._auto_scaling_group.getDesiredCapacity())+" to "+str(self._auto_scaling_group.getDesiredCapacity() - instancesDown)+" new desired capacity: "+str(self._auto_scaling_group.getDesiredCapacity() - instancesDown))
            #self._cooldown = True
            microservice._scale_down_trigger = 0
            #return True
        return True

    def arima_call(self, dataset, output, next_forecast, queue):
        timeseries = Timeseries(dataset)
        timeseries.execute(output, next_forecast)
        print()
        queue.put(timeseries)
    
    def update_file(self, scale):        
        workbook = load_workbook(filename = 'dataset\\'+self._auto_scaling_group.getAutoScalingGroupName()+'\\all.xlsx')
        worksheet = workbook['Sheet1']
        worksheet.cell(column=4,row=worksheet.max_row, value=scale)
        workbook.save(filename = 'dataset\\'+self._auto_scaling_group.getAutoScalingGroupName()+'\\all.xlsx')
        workbook.close()

    def save_file(self, cpu, network):
        workbook = load_workbook(filename = 'dataset\\'+self._auto_scaling_group.getAutoScalingGroupName()+'\\all.xlsx')
        worksheet = workbook['Sheet1']        
        
        worksheet.cell(column=5,row=worksheet.max_row, value=str(self._CPU_UPPER_TRESHOLD))        
        worksheet.cell(column=6,row=worksheet.max_row, value=str(self._CPU_LOWER_TRESHOLD))        
        worksheet.cell(column=7,row=worksheet.max_row, value=str(self._NETWORK_UPPER_TRESHOLD))       
        worksheet.cell(column=8,row=worksheet.max_row, value=str(self._NETWORK_LOWER_TRESHOLD))     

        try:
            worksheet.cell(column=9,row=worksheet.max_row, value=cpu._arima_order)
        except:
            pass
        try:
            worksheet.cell(column=10,row=worksheet.max_row, value=str(cpu._accuracy))
            worksheet.cell(column=11,row=worksheet.max_row, value=str(cpu._forecast[0]))
        except:
            pass        
        try:
            worksheet.cell(column=12,row=worksheet.max_row, value=network._arima_order)
        except:
            pass
        try:
            worksheet.cell(column=13,row=worksheet.max_row, value=str(network._accuracy))
            worksheet.cell(column=14,row=worksheet.max_row, value=str(network._forecast[0]))
        except:
            pass

        try:       
            worksheet.cell(column=15,row=worksheet.max_row, value=str(cpu._mape))
            worksheet.cell(column=16,row=worksheet.max_row, value=str(cpu._me))
            worksheet.cell(column=17,row=worksheet.max_row, value=str(cpu._mae))
            worksheet.cell(column=18,row=worksheet.max_row, value=str(cpu._mpe))
            worksheet.cell(column=19,row=worksheet.max_row, value=str(cpu._rmse))
            worksheet.cell(column=20,row=worksheet.max_row, value=str(cpu._corr))
            worksheet.cell(column=21,row=worksheet.max_row, value=str(cpu._minmax))
            worksheet.cell(column=22,row=worksheet.max_row, value=str(cpu._acf1))

            worksheet.cell(column=23,row=worksheet.max_row, value=str(network._mape))
            worksheet.cell(column=24,row=worksheet.max_row, value=str(network._me))
            worksheet.cell(column=25,row=worksheet.max_row, value=str(network._mae))
            worksheet.cell(column=26,row=worksheet.max_row, value=str(network._mpe))
            worksheet.cell(column=27,row=worksheet.max_row, value=str(network._rmse))
            worksheet.cell(column=28,row=worksheet.max_row, value=str(network._corr))
            worksheet.cell(column=29,row=worksheet.max_row, value=str(network._minmax))
            worksheet.cell(column=30,row=worksheet.max_row, value=str(network._acf1))
        except:
            pass

        worksheet.cell(column=31,row=worksheet.max_row, value=str(self._auto_scaling_group.getDesiredCapacity()))



        workbook.save(filename = 'dataset\\'+self._auto_scaling_group.getAutoScalingGroupName()+'\\all.xlsx')
        workbook.close()

    def scale(self, microservice):
        if microservice._cpu_utilization >= self._SCALE_UP:
            total =  (microservice._cpu_total * 100) / ((microservice._count+1) * 100) 
            if total >= self._SCALE_UP:
                self.update_file(1)
                return self.scale_up(2)
            else:
                self.update_file(1)
                return self.scale_up(1)
        elif microservice._cpu_utilization <= self._SCALE_DOWN and microservice._count > 1:            
            try:
                total =  (microservice._cpu_total * 100) / ((microservice._count-1) * 100)
            except:
                total =  (microservice._cpu_total * 100) / ((microservice._count) * 100)

            if total <= self._SCALE_DOWN and microservice._count > 2:
                self.update_file(2)
                return self.scale_down(2)
            else:
                self.update_file(2)
                return self.scale_down(1)
        
        return False

    def compare_proactive(self, value1, value2):
        if value1 > value2:
            return True
        return False
        
    def scale_v2(self, microservice, proactive):
        if proactive == True:
            if (microservice._cpu_utilization > self._CPU_UPPER_TRESHOLD and microservice._cpu_accuracy >= self._ACCURACY) or (microservice._network > self._NETWORK_UPPER_TRESHOLD and microservice._network_accuracy >= self._ACCURACY):
                '''
                total =  (microservice._cpu_total * 100) / ((microservice._count+1) * 100) 
                if total >= self._CPU_UPPER_TRESHOLD:
                    self.update_file("pro up")
                    print("PROATIVO")
                    microservice._scale_up_trigger += 1
                    print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                    return self.scale_up(2, microservice)
                else:
                '''
                self.update_file("pro up")
                print("PROATIVO")
                microservice._scale_up_trigger += 1
                print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                return self.scale_up(1, microservice)
            elif microservice._count > 1 and (microservice._cpu_utilization > self._CPU_UPPER_TRESHOLD and microservice._cpu_accuracy >= self._ACCURACY) or (microservice._network > self._NETWORK_UPPER_TRESHOLD and microservice._network_accuracy >= self._ACCURACY):      
                '''
                try:
                    total =  (microservice._cpu_total * 100) / ((microservice._count-1) * 100)
                except:
                    total =  (microservice._cpu_total * 100) / ((microservice._count) * 100)

                if total <= self._CPU_LOWER_TRESHOLD and microservice._count > 2:
                    self.update_file(2)
                    print("PROATIVO")
                    microservice._scale_down_trigger += 1
                    print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                    self.update_file("pro down")
                    return self.scale_down(2, microservice)
                else:
                '''
                self.update_file(2)
                print("PROATIVO")
                microservice._scale_down_trigger += 1
                print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                self.update_file("pro down")
                return self.scale_down(1, microservice)
        else:
            if microservice._cpu_utilization > self._CPU_UPPER_TRESHOLD or microservice._network > self._NETWORK_UPPER_TRESHOLD:
                '''
                total =  (microservice._cpu_total * 100) / ((microservice._count+1) * 100)
                if total >= self._CPU_UPPER_TRESHOLD:
                    self.update_file("rea up")
                    print("REATIVO")
                    microservice._scale_up_trigger += 1
                    print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                    return self.scale_up(2, microservice)
                else:
                '''
                self.update_file("rea up")
                print("REATIVO")
                microservice._scale_up_trigger += 1
                print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                return self.scale_up(1, microservice)
            elif microservice._count > 1 and (microservice._cpu_utilization < self._CPU_LOWER_TRESHOLD or microservice._network < self._NETWORK_LOWER_TRESHOLD):      
                '''
                try:
                    total =  (microservice._cpu_total * 100) / ((microservice._count-1) * 100)
                except:
                    total =  (microservice._cpu_total * 100) / ((microservice._count) * 100)

                if total <= self._CPU_LOWER_TRESHOLD and microservice._count > 2:
                    self.update_file("rea down")
                    print("REATIVO")
                    microservice._scale_down_trigger += 1
                    print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                    return self.scale_down(2, microservice)
                else:
                ''' 
                self.update_file("rea down")
                print("REATIVO")
                microservice._scale_down_trigger += 1
                print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                return self.scale_down(1, microservice)

        return False 

    def scale_cpu(self, microservice, proactive):
        if proactive == True:
            if (microservice._cpu_utilization > self._CPU_UPPER_TRESHOLD and microservice._cpu_accuracy >= self._ACCURACY):
                total =  (microservice._cpu_total * 100) / ((microservice._count+1) * 100) 
                if total >= self._CPU_UPPER_TRESHOLD:
                    self.update_file("pro up")
                    print("PROATIVO")
                    microservice._scale_up_trigger += 1
                    print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                    return self.scale_up(2, microservice)
                else:
                    self.update_file("pro up")
                    print("PROATIVO")
                    microservice._scale_up_trigger += 1
                    print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                    return self.scale_up(1, microservice)
            elif microservice._count > 1 and (microservice._cpu_utilization > self._CPU_UPPER_TRESHOLD and microservice._cpu_accuracy >= self._ACCURACY):      
                try:
                    total =  (microservice._cpu_total * 100) / ((microservice._count-1) * 100)
                except:
                    total =  (microservice._cpu_total * 100) / ((microservice._count) * 100)

                if total <= self._CPU_LOWER_TRESHOLD and microservice._count > 2:
                    self.update_file(2)
                    print("PROATIVO")
                    microservice._scale_down_trigger += 1
                    print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                    self.update_file("pro down")
                    return self.scale_down(2, microservice)
                else:
                    self.update_file(2)
                    print("PROATIVO")
                    microservice._scale_down_trigger += 1
                    print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                    self.update_file("pro down")
                    return self.scale_down(1, microservice)
        else:
            if microservice._cpu_utilization > self._CPU_UPPER_TRESHOLD:
                total =  (microservice._cpu_total * 100) / ((microservice._count+1) * 100)
                if total >= self._CPU_UPPER_TRESHOLD:
                    self.update_file("rea up")
                    print("REATIVO")
                    microservice._scale_up_trigger += 1
                    print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                    return self.scale_up(2, microservice)
                else:
                    self.update_file("rea up")
                    print("REATIVO")
                    microservice._scale_up_trigger += 1
                    print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                    return self.scale_up(1, microservice)
            elif microservice._count > 1 and (microservice._cpu_utilization < self._CPU_LOWER_TRESHOLD):      
                try:
                    total =  (microservice._cpu_total * 100) / ((microservice._count-1) * 100)
                except:
                    total =  (microservice._cpu_total * 100) / ((microservice._count) * 100)

                if total <= self._CPU_LOWER_TRESHOLD and microservice._count > 2:
                    self.update_file("rea down")
                    print("REATIVO")
                    microservice._scale_down_trigger += 1
                    print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                    return self.scale_down(2, microservice)
                else:
                    self.update_file("rea down")
                    print("REATIVO")
                    microservice._scale_down_trigger += 1
                    print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                    return self.scale_down(1, microservice)

        return False 

    def scale_network(self, microservice, proactive):
        if proactive == True:
            if (microservice._network > self._NETWORK_UPPER_TRESHOLD and microservice._network_accuracy >= self._ACCURACY):
                total =  (microservice._cpu_total * 100) / ((microservice._count+1) * 100) 
                if total >= self._CPU_UPPER_TRESHOLD:
                    self.update_file("pro up")
                    print("PROATIVO")
                    microservice._scale_up_trigger += 1
                    print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                    return self.scale_up(2, microservice)
                else:
                    self.update_file("pro up")
                    print("PROATIVO")
                    microservice._scale_up_trigger += 1
                    print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                    return self.scale_up(1, microservice)
            elif microservice._count > 1 and (microservice._network > self._NETWORK_UPPER_TRESHOLD and microservice._network_accuracy >= self._ACCURACY):      
                try:
                    total =  (microservice._cpu_total * 100) / ((microservice._count-1) * 100)
                except:
                    total =  (microservice._cpu_total * 100) / ((microservice._count) * 100)

                if total <= self._CPU_LOWER_TRESHOLD and microservice._count > 2:
                    self.update_file(2)
                    print("PROATIVO")
                    microservice._scale_down_trigger += 1
                    print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                    self.update_file("pro down")
                    return self.scale_down(2, microservice)
                else:
                    self.update_file(2)
                    print("PROATIVO")
                    microservice._scale_down_trigger += 1
                    print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                    self.update_file("pro down")
                    return self.scale_down(1, microservice)
        else:
            if microservice._network > self._NETWORK_UPPER_TRESHOLD:
                total =  (microservice._cpu_total * 100) / ((microservice._count+1) * 100)
                if total >= self._CPU_UPPER_TRESHOLD:
                    self.update_file("rea up")
                    print("REATIVO")
                    microservice._scale_up_trigger += 1
                    print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                    return self.scale_up(2, microservice)
                else:
                    self.update_file("rea up")
                    print("REATIVO")
                    microservice._scale_up_trigger += 1
                    print("Scale Up Trigger: "+str(microservice._scale_up_trigger))
                    return self.scale_up(1, microservice)
            elif microservice._count > 1 and (microservice._network < self._NETWORK_LOWER_TRESHOLD):      
                try:
                    total =  (microservice._cpu_total * 100) / ((microservice._count-1) * 100)
                except:
                    total =  (microservice._cpu_total * 100) / ((microservice._count) * 100)

                if total <= self._CPU_LOWER_TRESHOLD and microservice._count > 2:
                    self.update_file("rea down")
                    print("REATIVO")
                    microservice._scale_down_trigger += 1
                    print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                    return self.scale_down(2, microservice)
                else:
                    self.update_file("rea down")
                    print("REATIVO")
                    microservice._scale_down_trigger += 1
                    print("Scale Down Trigger: "+str(microservice._scale_down_trigger))
                    return self.scale_down(1, microservice)

        return False 
    def arima(self):
        q1 = queue.Queue()
        q2 = queue.Queue()

        t1 = threading.Thread(target=self.arima_call, args=(self._auto_scaling_group.getAutoScalingGroupName()+'\\cpu', False, 3, q1))
        t2 = threading.Thread(target=self.arima_call, args=(self._auto_scaling_group.getAutoScalingGroupName()+'\\network', False, 3, q2))
    
        # starting thread
        t1.start()
        t2.start()
    
        # wait until thread is completely executed
        t1.join()
        t2.join()
  
        return q1.get(), q2.get()

    def proactive_scale(self, microservice):
        # creating thread
        print("Start Forecasting")

        now = datetime.datetime.now()

        cpu, network = self.arima()

        now = datetime.datetime.now() - now

        self.save_file(cpu, network)

        try:
            microservice._cpu_utilization = cpu._forecast[0]
        except:
            pass
        try:
            microservice._network = network._forecast[0]
        except:
            pass

        microservice._cpu_accuracy = cpu._accuracy
        microservice._network_accuracy = network._accuracy
        
        #return self.scale(microservice)
        return self.scale_v2(microservice, True)

        #se precisão de cpu e rede forem >= a 70% eo valor atual não for praticamente ZERO				
	        #se o próximo valor é maior que antigo			
		        #se o próximo valor for muito maior que o antigo		
			        #aumenta uma instnacia	
				        #cooldown de quatro lidas
        '''
        try:
            cpu_count = 0
            NETWORK_count = 0
            netout_count = 0

            # crescimento de CPU é diretamente proporcional ao crescimento de Rede




            for instance in microservice._instances:
                if abs(float(cpu._forecast[0])-float(microservice._cpu_total)) >= self._PROACTIVE_DIFF:
                    cpu_count +=1
                if abs(float(netin._forecast[0])-float(microservice._network_in)) >= self._PROACTIVE_DIFF:
                    NETWORK_count +=1
                if abs(float(netout._forecast[0])-float(microservice._network_out)) >= self._PROACTIVE_DIFF:
                    netout_count +=1
            
            if float(cpu._accuracy) >= self._ACCURACY or float(netin._accuracy) >= self._ACCURACY or float(netout._accuracy) >= self._ACCURACY:
                if cpu_count > 0 or NETWORK_count > 0 or netout_count > 0:
                    print("PROATIVO")
                    return self.scale(microservice)
        except Exception as e:
            print(e)
        '''
        print(str(now))

        return False

    def reactive_scale(self, microservice):
        #if not self.scale(microservice):
        if not self.scale_v2(microservice, False):
            self.update_file(0)
            return False
        return True

    def aws(self):
        self.save_file("", "")

    def execute(self, microservice):
        if not self.reactive_scale(microservice):
            return self.proactive_scale(microservice)
        else:            
            self.save_file("", "")
            return True