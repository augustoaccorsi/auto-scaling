from enum import auto
import json, os, xlsxwriter, datetime, timedelta, boto3, sys
from openpyxl import load_workbook
from autoscalinggroup import AutoScalingGroup, Instance, AvailabilityZone, LoadBalancer, EnabledMetric, Tags
from autoscaling import Autoscaling
from microservice import Microservice
import weakref

class App():
    _alive = []
    #def __init__(self, autoscalinggroup, region, accessKeyId, secretAccessKey, sessionToken):
    def __new__(self, autoscalinggroup, region):
        
        self = super().__new__(self)
        App._alive.append(self)

        self._autoscalinggroup = autoscalinggroup
        self._asg = boto3.client('autoscaling', region_name=region)
        self._cloud_watch = boto3.client('cloudwatch',  region_name=region)
        self._ec2 = boto3.client('ec2', region_name=region)
        self._elb = boto3.client('elb', region_name=region)

        self._region = region

        self._up = False

        self.lb_instnaces = []
        self._lb_name = ""

        #self._asg = boto3.client('autoscaling', region_name=region, aws_access_key_id=accessKeyId, aws_secret_access_key=secretAccessKey, aws_session_token=sessionToken)
        #self._cloud_watch = boto3.client('cloudwatch',  region_name=region, aws_access_key_id=accessKeyId, aws_secret_access_key=secretAccessKey, aws_session_token=sessionToken)

        self._auto_scaling_info = dict()
        self._auto_scaling_group = AutoScalingGroup()
        self._instances = []
        self._instances_ids = dict()
        
        self._cooldown = False
        self._cooldown_trigger = 0
        self._timeseries = False
        self._COOLDOWN = 3

        self._microservice = Microservice(self._instances)

        self.describe()
        
        return weakref.proxy(self)
    
    def commit_suicide(self):
        self._alive.remove(self)
    
    def renew_connection(self):
        self._asg = boto3.client('autoscaling', region_name=self._region)
        self._cloud_watch = boto3.client('cloudwatch',  region_name=self._region)
        self._ec2 = boto3.client('ec2',  region_name=self._region)

    def clear_all(self):
        self._auto_scaling_info = dict()
        self._auto_scaling_group = AutoScalingGroup()
        #self._instances = []
        #self._instances_ids = dict()

    def describe(self):
        
        if len(self._instances) > 0:
            self.clear_all()
            self.renew_connection()
        
        self._auto_scaling_info = self._asg.describe_auto_scaling_groups(AutoScalingGroupNames=[self._autoscalinggroup])
        self.build_auto_scaling_group()
        return self

    def build_load_balancers(self, loadBalancers):
        loadBalancersList = []
        for lb in loadBalancers:
            ldb = LoadBalancer()
            ldb.setLoadBalancer(lb)
            loadBalancersList.append(ldb)
            loadBalancer = (self._elb.describe_load_balancers(LoadBalancerNames=[lb]))

            if not loadBalancer['LoadBalancerDescriptions'][0]['Instances'] in self._instances:
                self._instances_ids = (loadBalancer['LoadBalancerDescriptions'][0]['Instances'])
                self._auto_scaling_group.appendInstanceId((loadBalancer['LoadBalancerDescriptions'][0]['Instances']))
                self._lb_name = (loadBalancer['LoadBalancerDescriptions'][0]['LoadBalancerName'])
            
        return loadBalancersList
    
    def build_instances(self, instances):
        
        instancesList = []
        for instance in self._instances_ids:      

            if instance['InstanceId'] not in self.lb_instnaces:
                
                self.lb_instnaces.append(instance['InstanceId'])

                inst = Instance()
                inst.setInstanceId(instance['InstanceId'])

                state = self._ec2.describe_instances(InstanceIds=[instance['InstanceId']])
                inst.setLifecycleState(state['Reservations'][0]['Instances'][0]['State']['Name'].title())
                inst.setLaunchTime(state['Reservations'][0]['Instances'][0]['LaunchTime'])

                health = self._elb.describe_instance_health(LoadBalancerName=self._lb_name, Instances=[{'InstanceId': instance['InstanceId']}])
                inst.setHealthStatus(health['InstanceStates'][0]['State'])

                try:
                    status = self._ec2.describe_instance_status(InstanceIds=[instance['InstanceId']])
                    inst.setStatus(status['InstanceStatuses'][0]['InstanceStatus']['Details'][0]['Status'].title())
                except:
                    inst.setStatus("-")

                instancesList.append(inst)
                self._instances.append(inst)

        self.update_instances()

        return self._instances

    def update_instances(self):
        for instance in self._instances:
            
            state = self._ec2.describe_instances(InstanceIds=[instance.getInstanceId()])
            instance.setLifecycleState(state['Reservations'][0]['Instances'][0]['State']['Name'].title())
            instance.setLaunchTime(state['Reservations'][0]['Instances'][0]['LaunchTime'])

            health = self._elb.describe_instance_health(LoadBalancerName=self._lb_name, Instances=[{'InstanceId': instance.getInstanceId()}])
            instance.setHealthStatus(health['InstanceStates'][0]['State'])

            try:
                status = self._ec2.describe_instance_status(InstanceIds=instance.getInstanceId())
                instance.setStatus(status['InstanceStatuses'][0]['InstanceStatus']['Details'][0]['Status'].title())
            except:
                instance.setStatus("-")
        return self._instances

    def build_auto_scaling_group(self):
        self._auto_scaling_group.setAutoScalingGroupName(self._auto_scaling_info['AutoScalingGroups'][0]['AutoScalingGroupName'])
        self._auto_scaling_group.setDesiredCapacity(self._auto_scaling_info['AutoScalingGroups'][0]['DesiredCapacity'])
        self._auto_scaling_group.setLoadBalancerNames(self.build_load_balancers(self._auto_scaling_info['AutoScalingGroups'][0]['LoadBalancerNames']))
        self._auto_scaling_group.setInstances(self.build_instances(self._auto_scaling_info['AutoScalingGroups'][0]['Instances']))

        self.create_files()

    def create_files(self):
        if not os.path.isfile('dataset\\'+self._autoscalinggroup+'\\cpu.xlsx'):
            workbook = xlsxwriter.Workbook('dataset\\'+self._autoscalinggroup+'\\cpu.xlsx')
            worksheet = workbook.add_worksheet()
            format = workbook.add_format({'num_format': 'dd/mm/yy hh:mm'})
            worksheet.write('A1', 'date', format)
            worksheet.write('B1', 'value')
            workbook.close()

        
        if not os.path.isfile('dataset\\'+self._autoscalinggroup+'\\network.xlsx'):
            workbook = xlsxwriter.Workbook('dataset\\'+self._autoscalinggroup+'\\network.xlsx')
            worksheet = workbook.add_worksheet()
            format = workbook.add_format({'num_format': 'dd/mm/yy hh:mm'})
            worksheet.write('A1', 'date', format)
            worksheet.write('B1', 'value')
            workbook.close()

        if not os.path.isfile('dataset\\'+self._autoscalinggroup+'\\all.xlsx'):
            workbook = xlsxwriter.Workbook('dataset\\'+self._autoscalinggroup+'\\all.xlsx')
            worksheet = workbook.add_worksheet()
            format = workbook.add_format({'num_format': 'dd/mm/yy hh:mm'})
            worksheet.write('A1', 'date', format) #1
            worksheet.write('B1', 'cpu') #2 
            worksheet.write('C1', 'net') #3
            worksheet.write('D1', 'scale') #4

            worksheet.write('E1', 'up_cpu') #5
            worksheet.write('F1', 'down_cpu') #6
            worksheet.write('G1', 'net_up') #7
            worksheet.write('H1', 'net_down') #8

            worksheet.write('I1', 'cpu_arima') #9           
            worksheet.write('J1', 'cpu_acc') #10
            worksheet.write('K1', 'cpu_pred') #11
            worksheet.write('L1', 'net_arima') #12        
            worksheet.write('M1', 'net_acc') #13
            worksheet.write('N1', 'net_pred') #14

            worksheet.write('O1', 'cpu_mape') #15
            worksheet.write('P1', 'cpu_me') #16
            worksheet.write('Q1', 'cpu_mae') #17
            worksheet.write('R1', 'cpu_mpe') #18
            worksheet.write('S1', 'cpu_rmse') #19
            worksheet.write('T1', 'cpu_corr') #20
            worksheet.write('U1', 'cpu_minmax') #21
            worksheet.write('V1', 'cpu_acf1') #22
            
            worksheet.write('W1', 'net_mape') #23
            worksheet.write('X1', 'net_me') #24
            worksheet.write('Y1', 'net_mae') #25
            worksheet.write('Z1', 'net_mpe') #26
            worksheet.write('AA1', 'net_rmse') #27
            worksheet.write('AB1', 'net_corr') #28
            worksheet.write('AC1', 'net_minmax') #29
            worksheet.write('AD1', 'net_acf1') #30
            worksheet.write('AE1', 'instancias') #31

            workbook.close()

    def save_into_file(self, datetime, microservice):

        workbook = load_workbook(filename = 'dataset\\'+self._autoscalinggroup+'\\cpu.xlsx')
        worksheet = workbook['Sheet1']
        
        newRowLocation = worksheet.max_row +1

        worksheet.cell(column=1,row=newRowLocation, value=datetime)
        if microservice._cpu_utilization != None:
            worksheet.cell(column=2,row=newRowLocation, value=microservice._cpu_utilization)

        workbook.save(filename = 'dataset\\'+self._autoscalinggroup+'\\cpu.xlsx')
        workbook.close()


        workbook = load_workbook(filename = 'dataset\\'+self._autoscalinggroup+'\\network.xlsx')
        worksheet = workbook['Sheet1']
        
        newRowLocation = worksheet.max_row +1

        worksheet.cell(column=1,row=newRowLocation, value=datetime)
        if microservice._network != None:
            worksheet.cell(column=2,row=newRowLocation, value=microservice._network)

        workbook.save(filename = 'dataset\\'+self._autoscalinggroup+'\\network.xlsx')
        workbook.close()

        workbook = load_workbook(filename = 'dataset\\'+self._autoscalinggroup+'\\all.xlsx')
        worksheet = workbook['Sheet1']
        
        newRowLocation = worksheet.max_row +1

        worksheet.cell(column=1,row=newRowLocation, value=datetime)
        if microservice._cpu_utilization != None:
            worksheet.cell(column=2,row=newRowLocation, value=microservice._cpu_utilization)
        if microservice._network != None:
            worksheet.cell(column=3,row=newRowLocation, value=microservice._network)

        workbook.save(filename = 'dataset\\'+self._autoscalinggroup+'\\all.xlsx')
        workbook.close()

    def get_metric(self, metric, instance, start_time, end_time, statistics):
        return self._cloud_watch.get_metric_statistics(Namespace='AWS/EC2',
            MetricName=metric,
            Dimensions=[
                {
                    'Name': 'InstanceId',
                    'Value': instance
                },
            ], StartTime=start_time, EndTime=end_time, Period=120, Statistics=[statistics]) 

    def read_instances(self):
        count = 0

        end_time = datetime.datetime.utcnow()
        
        start_time = end_time - datetime.timedelta(seconds=120) #buscar dados dos ultimos 2 minutos

        start_time = start_time.strftime('%m/%d/%Y %H:%M:%S')
        end_time = end_time.strftime('%m/%d/%Y %H:%M:%S')

        for instance in self._instances:

            if instance.getLifecycleState() == "Terminated":
                self._instances.remove(instance)
                break
            else:

                count+=1
            
                cpu =  self.get_metric("CPUUtilization", instance.getInstanceId(), start_time, end_time, "Average")
                networkIn =  self.get_metric("NetworkIn", instance.getInstanceId(), start_time, end_time, "Average")
                networkOut =  self.get_metric("NetworkOut", instance.getInstanceId(), start_time, end_time, "Average")
                        
                state = self._ec2.describe_instances(InstanceIds=[instance.getInstanceId()])
                instance.setLifecycleState(state['Reservations'][0]['Instances'][0]['State']['Name'].title())
                instance.setLaunchTime(state['Reservations'][0]['Instances'][0]['LaunchTime'])
                try:
                    status = self._ec2.describe_instance_status(InstanceIds=[instance.getInstanceId()])
                    instance.setStatus(status['InstanceStatuses'][0]['InstanceStatus']['Details'][0]['Status'].title())
                except:
                    instance.setStatus("-")

                print("Instance "+str(count)+":  "+instance.getInstanceId())
                print("Lifecycle State: "+instance.getLifecycleState()+" - "+instance.getHealthStatus()+" - "+instance.getStatus())
        
                try:
                    cpuUtilization = round(float(cpu['Datapoints'][0]['Average']),4)
                    print("CPU Usage: "+str(cpuUtilization)+"%")
                    instance.setCpuUtilization(cpuUtilization)
                except:
                    cpuUtilization = None
                
                try:
                    netIn = str(float(networkIn['Datapoints'][0]['Average'])/1000) # valor em kB
                    #print("Network In: "+netIn+"Kb")
                    instance.setNetworkIn(netIn)
                except:
                    netIn = None
                
                try:
                    netOut = str(float(networkOut['Datapoints'][0]['Average'])/1000) # valor em kB
                    #print("Network Out: "+netOut+"Kb")
                    instance.setNetworkOut(netOut)
                except:
                    netOut = None
                try:
                    network = float(instance.getNetworkOut()) + float(instance.getNetworkOut())
                    print("Network: "+str(network)+"Kb")
                    instance.setNetwork(network)
                except:
                    network = None

                print()

        if self._cooldown == False:
            
            self._microservice.set_all(self._instances)
            
            print("----")
            #print("Total CPU: "+ str(self._microservice._cpu_total))
            print("Utilization: "+str(self._microservice._cpu_utilization))
            print("Network: "+ str(self._microservice._network))
            print("----")
            print()

            if self._microservice._cpu_total > 0 and self._microservice._cpu_utilization > 0 and self._microservice._network > 0:
                self.save_into_file(end_time, self._microservice)
            
                autoscaling = Autoscaling(self._instances, self._auto_scaling_group, self._asg)
                
                autoscaling.aws()
                
                
                '''
                if not autoscaling.execute(self._microservice):
                    self._microservice._scale_up_trigger = 0
                    self._microservice._scale_down_trigger = 0
                           
                self._cooldown = autoscaling._cooldown
                '''
        else:
            self._cooldown_trigger +=1
            print("Cooldown "+str(self._cooldown_trigger))
            if self._cooldown_trigger == self._COOLDOWN:
                self._cooldown = False
                self._cooldown_trigger = 0
        self.describe()

    def scale_up(self):
        self._asg.set_desired_capacity(AutoScalingGroupName=self._auto_scaling_group.getAutoScalingGroupName(), DesiredCapacity=(self._auto_scaling_group.getDesiredCapacity() + 1))
        print("Autoscaling Group scalled up, from "+str(self._auto_scaling_group.getDesiredCapacity())+" to "+str(self._auto_scaling_group.getDesiredCapacity() + 1)+" new desired capacity: "+str(self._auto_scaling_group.getDesiredCapacity() + 1))
    
    def scale_down(self):
        self._asg.set_desired_capacity(AutoScalingGroupName=self._auto_scaling_group.getAutoScalingGroupName(), DesiredCapacity=(self._auto_scaling_group.getDesiredCapacity() - 1))
        print("Autoscaling Group scalled down, from "+str(self._auto_scaling_group.getDesiredCapacity())+" to "+str(self._auto_scaling_group.getDesiredCapacity() - 1)+" new desired capacity: "+str(self._auto_scaling_group.getDesiredCapacity() - 1))


if __name__ == '__main__':
    app = App(sys.argv[1], "sa-east-1")
    
    try:
        if sys.argv[2] == "up":
            app.scale_up()
        if sys.argv[2] == "down":
            app.scale_down()
    except:
        app.create_files()
        app.read_instances()